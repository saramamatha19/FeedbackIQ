"""Stage 1 (Reading File) and Stage 2 (Cleaning Data) helpers, shared by the
CSV, Excel, and paste ingestion paths."""

import csv
import io
import re

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024
MIN_TEXT_LENGTH = 1
MAX_TEXT_LENGTH = 5000

_CONTROL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
_WHITESPACE_RE = re.compile(r"\s+")


class InvalidFileError(Exception):
    pass


_HEADER_HINTS = ("feedback", "review", "comment", "text", "description")
_HEADER_EXCLUDE = ("id", "date", "time", "type", "category", "rating", "score", "severity", "status", "count", "name")


def _average_lengths(rows: list[list], num_cols: int) -> list[float]:
    """Average non-empty stringified value length per column, across sampled rows."""
    sums = [0] * num_cols
    counts = [0] * num_cols
    for row in rows:
        for i in range(num_cols):
            value = row[i] if i < len(row) else None
            if value is None:
                continue
            length = len(str(value).strip())
            if length == 0:
                continue
            sums[i] += length
            counts[i] += 1
    return [(sums[i] / counts[i] if counts[i] else 0.0) for i in range(num_cols)]


def _select_feedback_column(headers: list[str], sample_rows: list[list]) -> int:
    """Picks the column most likely to hold free-form feedback/review text.

    Priority: an exact 'feedback' header always wins (explicit intent). Otherwise, prefer
    columns whose header hints at review/feedback content (excluding metadata-ish headers
    like 'feedback_type' or 'review_id'), breaking ties by average content length — real
    review text is reliably much longer than IDs, dates, categories, or ratings. Falls back
    to the single longest-content column, then the first column, if no header hints match.
    """
    normalized = [h.strip().lower() for h in headers]

    for i, header in enumerate(normalized):
        if header == "feedback":
            return i

    avg_lengths = _average_lengths(sample_rows, len(headers))

    hinted = [
        i
        for i, header in enumerate(normalized)
        if any(hint in header for hint in _HEADER_HINTS) and not any(bad in header for bad in _HEADER_EXCLUDE)
    ]
    if hinted:
        return max(hinted, key=lambda i: avg_lengths[i])

    if any(avg_lengths):
        return max(range(len(headers)), key=lambda i: avg_lengths[i])

    return 0


def clean_text(raw_text: str) -> str | None:
    """Returns cleaned text, or None if the row should be dropped (empty/garbage)."""
    text = _CONTROL_CHARS_RE.sub("", raw_text)
    text = _WHITESPACE_RE.sub(" ", text).strip()
    if len(text) < MIN_TEXT_LENGTH:
        return None
    if len(text) > MAX_TEXT_LENGTH:
        text = text[:MAX_TEXT_LENGTH]
    return text


def parse_csv(file_bytes: bytes) -> list[str]:
    """Extracts feedback text from a CSV, auto-detecting which column holds the
    review/feedback content (see _select_feedback_column)."""
    if len(file_bytes) > MAX_FILE_SIZE_BYTES:
        raise InvalidFileError("File exceeds the 5MB size limit.")
    if not file_bytes.strip():
        raise InvalidFileError("The uploaded file is empty.")

    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = file_bytes.decode("latin-1")
        except UnicodeDecodeError as exc:
            raise InvalidFileError("Could not decode file — please upload a UTF-8 CSV.") from exc

    try:
        all_rows = list(csv.reader(io.StringIO(text)))
    except csv.Error as exc:
        raise InvalidFileError(f"Could not parse CSV: {exc}") from exc

    if not all_rows:
        raise InvalidFileError("CSV has no header row.")

    headers, data_rows = all_rows[0], all_rows[1:]
    feedback_col = _select_feedback_column(headers, data_rows)

    rows = []
    for row in data_rows:
        if feedback_col >= len(row):
            continue
        value = row[feedback_col].strip()
        if value:
            rows.append(value)

    if not rows:
        raise InvalidFileError(f"No feedback text found in column '{headers[feedback_col]}'.")
    return rows


def parse_xlsx(file_bytes: bytes) -> list[str]:
    """Extracts feedback text from an Excel workbook's first sheet, auto-detecting which
    column holds the review/feedback content (see _select_feedback_column)."""
    if len(file_bytes) > MAX_FILE_SIZE_BYTES:
        raise InvalidFileError("File exceeds the 5MB size limit.")
    if not file_bytes.strip():
        raise InvalidFileError("The uploaded file is empty.")

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise InvalidFileError(f"Could not parse Excel file: {exc}") from exc

    sheet = workbook.active
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        raise InvalidFileError("Excel file has no header row.")
    header_row, data_rows = all_rows[0], all_rows[1:]

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    feedback_col = _select_feedback_column(headers, data_rows)

    rows = []
    for row in data_rows:
        if feedback_col >= len(row) or row[feedback_col] is None:
            continue
        value = str(row[feedback_col]).strip()
        if value:
            rows.append(value)
    workbook.close()

    if not rows:
        raise InvalidFileError(f"No feedback text found in column '{headers[feedback_col]}'.")
    return rows
